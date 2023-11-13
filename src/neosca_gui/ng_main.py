#!/usr/bin/env python3

import copy
import glob
import os
import os.path as os_path
import re
import subprocess
import sys
import textwrap
from typing import Any, Dict, Generator, Iterable, List, Literal, Optional, Set, Union

from PySide6.QtCore import QModelIndex, QObject, Qt, QThread, Signal
from PySide6.QtGui import (
    QAction,
    QCursor,
    QPalette,
    QStandardItem,
    QStandardItemModel,
)
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QCheckBox,
    QComboBox,
    QDialog,
    QFileDialog,
    QFontDialog,
    QGridLayout,
    QGroupBox,
    QHeaderView,
    QLabel,
    QMainWindow,
    QMenu,
    QMessageBox,
    QPushButton,
    QRadioButton,
    QScrollArea,
    QSplitter,
    QTableView,
    QTabWidget,
    QTextEdit,
    QWidget,
)

from .neosca.lca.lca import LCA
from .neosca.neosca import NeoSCA
from .neosca.structure_counter import StructureCounter
from .ng_about import __title__, __version__
from .ng_io import SCAIO
from .ng_settings_default import (
    DEFAULT_FONT_FAMILY,
    DEFAULT_FONT_SIZE,
    DEFAULT_INTERFACE_SCALING,
    settings_default,
)


class Ng_QSS_Loader:
    def __init__(self):
        pass

    @staticmethod
    def read_qss_file(qss_file_path: str, default: Any = ""):
        if os_path.isfile(qss_file_path) and os_path.getsize(qss_file_path) > 0:
            with open(qss_file_path, encoding="utf-8") as file:
                return file.read()
        else:
            return default

    @staticmethod
    def get_qss_value(qss: str, selector: str, attrname: str) -> Optional[str]:
        """
        >>> qss = "QHeaderView::section:horizontal { background-color: #5C88C5; }"
        >>> get_qss_value(qss, "QHeaderView::section:horizontal", "background-color")
        5C88C5
        """
        # Notice that only the 1st selector will be matched here
        matched_selector = re.search(selector, qss)
        if matched_selector is None:
            return None
        matched_value = re.search(rf"[^}}]+{attrname}:\s*([^;]+);", qss[matched_selector.end() :])
        if matched_value is None:
            return None
        return matched_value.group(1)


class Ng_Model(QStandardItemModel):
    data_cleared = Signal()
    data_changed = Signal()
    data_exported = Signal()

    def __init__(self, *args, orientation: Literal["hor", "ver"] = "hor", **kwargs) -> None:
        super().__init__(*args, **kwargs)
        self.orientation = orientation

        self.has_been_exported: bool = False
        self.data_exported.connect(lambda: self.set_has_been_exported(True))
        self.data_changed.connect(lambda: self.set_has_been_exported(False))

    def set_item_str(self, rowno: int, colno: int, value: str) -> None:
        item = QStandardItem(value)
        item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        self.setItem(rowno, colno, item)

    def set_row_str(self, rowno: int, values: Iterable[str]) -> None:
        for colno, value in enumerate(values):
            self.set_item_str(rowno, colno, value)

    def set_item_num(self, rowno: int, colno: int, value: Union[int, float, str]) -> None:
        if not isinstance(value, str):
            value = str(value)
        item = QStandardItem(value)
        item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.setItem(rowno, colno, item)

    def set_row_num(self, rowno: int, values: Iterable[Union[int, float, str]]) -> None:
        for colno, value in enumerate(values):
            self.set_item_num(rowno, colno, value)

    def set_has_been_exported(self, exported: bool) -> None:
        self.has_been_exported = exported

    def set_single_empty_row(self) -> None:
        # https://stackoverflow.com/questions/75038194/qt6-how-to-disable-selection-for-empty-cells-in-qtableview
        if self.orientation == "hor":
            self.setRowCount(0)
            self.setRowCount(1)
        elif self.orientation == "ver":
            self.setColumnCount(0)
            self.setColumnCount(1)
        self.data_cleared.emit()

    def remove_single_empty_row(self) -> None:
        if self.rowCount() == 1 and self.item(0, 0) is None:
            self.setRowCount(0)


class Ng_TableView(QTableView):
    def __init__(
        self,
        *args,
        main,
        model: Ng_Model,
        has_horizontal_header: bool = True,
        has_vertical_header: bool = True,
        **kwargs,
    ) -> None:
        super().__init__(*args, **kwargs)
        self.main = main
        self.setModel(model)
        self.model_: Ng_Model = model
        self.model_.data_changed.connect(self.after_data_changed)
        self.model_.data_cleared.connect(self.after_data_changed)
        self.has_horizontal_header = has_horizontal_header
        self.has_vertical_header = has_vertical_header

        self.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.setFocusPolicy(Qt.FocusPolicy.ClickFocus)
        self.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.setSelectionBehavior(QTableView.SelectionBehavior.SelectRows)
        self.horizontalHeader().setHighlightSections(False)
        self.verticalHeader().setHighlightSections(False)
        self.after_data_changed()

    def after_data_changed(self) -> None:
        self.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)

    def model(self) -> Ng_Model:
        """Override QTableView().model()"""
        return self.model_

    def set_openpyxl_horizontal_header_alignment(self, cell) -> None:
        from openpyxl.styles import Alignment

        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def set_openpyxl_vertical_header_alignment(self, cell) -> None:
        from openpyxl.styles import Alignment

        if self.has_vertical_header:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def set_openpyxl_cell_alignment(self, cell, item: QStandardItem) -> None:
        # https://doc.qt.io/qtforpython-6/PySide6/QtCore/Qt.html#PySide6.QtCore.PySide6.QtCore.Qt.AlignmentFlag
        # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.styles.alignment.html
        # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L887
        # Qt
        #  Horizontal: Qt.AlignLeft, Qt.AlignRight,  Qt.AlignHCenter, Qt.AlignJustify
        #  Vertical:   Qt.AlignTop,  Qt.AlignBottom, Qt.AlignVCenter, Qt.AlignBaseline
        # OpenPyXL
        #  Horizontal: justify, center, distributed, left, right, fill, general, centerContinuous
        #  Vertical:   justify, center, distributed, top,  bottom

        from openpyxl.styles import Alignment

        alignment_item: Qt.AlignmentFlag = item.textAlignment()

        # Horizontal
        if Qt.AlignmentFlag.AlignLeft in alignment_item:
            alignment_cell_horizontal = "left"
        elif Qt.AlignmentFlag.AlignRight in alignment_item:
            alignment_cell_horizontal = "right"
        elif Qt.AlignmentFlag.AlignHCenter in alignment_item:
            alignment_cell_horizontal = "center"
        elif Qt.AlignmentFlag.AlignJustify in alignment_item:
            alignment_cell_horizontal = "justify"
        else:
            alignment_cell_horizontal = "left"

        # Vertical
        if Qt.AlignmentFlag.AlignTop in alignment_item:
            alignment_cell_vertical = "top"
        elif Qt.AlignmentFlag.AlignBottom in alignment_item:
            alignment_cell_vertical = "bottom"
        elif Qt.AlignmentFlag.AlignVCenter in alignment_item:
            alignment_cell_vertical = "center"
        # > Wordless: Not sure
        elif Qt.AlignmentFlag.AlignBaseline in alignment_item:
            alignment_cell_vertical = "justify"
        else:
            alignment_cell_vertical = "center"

        cell.alignment = Alignment(
            horizontal=alignment_cell_horizontal, vertical=alignment_cell_vertical, wrap_text=True
        )

    def export_table(self) -> None:
        file_path, file_type = QFileDialog.getSaveFileName(
            parent=None,
            caption="Export Table",
            dir=os_path.normpath(os_path.expanduser("~/Desktop")),
            filter="Excel Workbook (*.xlsx);;CSV File (*.csv);;TSV File (*.tsv)",
        )
        if not file_path:
            return

        model: Ng_Model = self.model()
        col_count = model.columnCount()
        row_count = model.rowCount()
        try:
            if ".xlsx" in file_type:
                # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L701C1-L716C54
                import openpyxl
                from openpyxl.styles import Font, PatternFill
                from openpyxl.utils import get_column_letter

                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                worksheet_cell = worksheet.cell

                rowno_cell_offset = 2 if self.has_horizontal_header else 1
                colno_cell_offset = 2 if self.has_vertical_header else 1

                # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L628C3-L629C82
                dpi_horizontal = QApplication.primaryScreen().logicalDotsPerInchX()
                dpi_vertical = QApplication.primaryScreen().logicalDotsPerInchY()

                # 1. Horizontal header text and alignment
                if self.has_horizontal_header:
                    for colno_cell, colno_item in enumerate(range(col_count)):
                        cell = worksheet_cell(1, colno_cell_offset + colno_cell)
                        cell.value = model.horizontalHeaderItem(colno_item).text()
                        self.set_openpyxl_horizontal_header_alignment(cell)
                # 2. Vertical header text and alignment
                if self.has_vertical_header:
                    for rowno_cell, rowno_item in enumerate(range(row_count)):
                        cell = worksheet_cell(rowno_cell_offset + rowno_cell, 1)
                        cell.value = model.verticalHeaderItem(rowno_item).text()
                        self.set_openpyxl_vertical_header_alignment(cell)

                # 3. Both header background and font
                # 3.0.1 Get header background
                horizon_bacolor: Optional[str] = Ng_QSS_Loader.get_qss_value(
                    self.main.styleSheet(), "QHeaderView::section:horizontal", "background-color"
                )
                vertical_bacolor: Optional[str] = Ng_QSS_Loader.get_qss_value(
                    self.main.styleSheet(), "QHeaderView::section:vertical", "background-color"
                )
                # 3.0.2 Get header font, currently only consider color and boldness
                #  https://www.codespeedy.com/change-font-color-of-excel-cells-using-openpyxl-in-python/
                #  https://doc.qt.io/qt-6/stylesheet-reference.html#font-weight
                font_color = Ng_QSS_Loader.get_qss_value(self.main.styleSheet(), "QHeaderView::section", "color")
                font_color = font_color.lstrip("#") if font_color is not None else "000"
                font_weight = Ng_QSS_Loader.get_qss_value(
                    self.main.styleSheet(), "QHeaderView::section", "font-weight"
                )
                is_bold = (font_weight == "bold") if font_weight is not None else False
                # 3.1 Horizontal header background and font
                if self.has_horizontal_header:
                    # 3.1.1 Horizontal header background
                    #  TODO: Currently all tabs share the same style sheet and the
                    #   single QSS file is loaded from MainWindow, thus here the
                    #   style sheet is accessed from self. In the future different
                    #   tabs might load their own QSS files, and the style sheet
                    #   should be accessed from the QTabWidget. This is also the
                    #   case for all other "self.styleSheet()" expressions, when
                    #   making this change, remember to edit all of them.
                    if horizon_bacolor is not None:
                        horizon_bacolor = horizon_bacolor.lstrip("#")
                        for colno in range(col_count):
                            cell = worksheet_cell(1, colno_cell_offset + colno)
                            cell.fill = PatternFill(fill_type="solid", fgColor=horizon_bacolor)
                    # 3.1.2 Horizontal header font
                    for colno in range(col_count):
                        cell = worksheet_cell(1, colno_cell_offset + colno)
                        cell.font = Font(color=font_color, bold=is_bold)
                # 3.2 Vertical header background and font
                if self.has_vertical_header:
                    # 3.2.1 Vertial header background
                    if vertical_bacolor is not None:
                        vertical_bacolor = vertical_bacolor.lstrip("#")
                        for rowno in range(row_count):
                            cell = worksheet_cell(rowno_cell_offset + rowno, 1)
                            cell.fill = PatternFill(fill_type="solid", fgColor=vertical_bacolor)
                    # 3.2.2 Vertical header font
                    for rowno in range(row_count):
                        cell = worksheet_cell(rowno_cell_offset + rowno, 1)
                        cell.font = Font(color=font_color, bold=is_bold)

                # 4. Cells
                for rowno in range(row_count):
                    for colno in range(col_count):
                        cell = worksheet_cell(rowno_cell_offset + rowno, colno_cell_offset + colno)
                        item = model.item(rowno, colno)
                        item_value = item.text()
                        try:  # noqa: SIM105
                            item_value = float(item_value)
                        except ValueError:
                            pass
                        cell.value = item_value
                        self.set_openpyxl_cell_alignment(cell, item)
                # 5. Column width
                for colno in range(col_count):
                    # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L729
                    worksheet.column_dimensions[get_column_letter(colno_cell_offset + colno)].width = (
                        self.horizontalHeader().sectionSize(colno) / dpi_horizontal * 13 + 3
                    )

                if self.has_vertical_header:
                    # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L731
                    worksheet.column_dimensions[get_column_letter(1)].width = (
                        self.verticalHeader().width() / dpi_horizontal * 13 + 3
                    )
                # 6. Row height
                worksheet.row_dimensions[1].height = self.horizontalHeader().height() / dpi_vertical * 72
                # 7. Freeze panes
                # https://stackoverflow.com/questions/73837417/freeze-panes-first-two-rows-and-column-with-openpyxl
                # Using "2" in both cases means to always freeze the 1st column
                if self.has_horizontal_header:
                    worksheet.freeze_panes = "B2"
                else:
                    worksheet.freeze_panes = "A2"
                workbook.save(file_path)
            elif ".csv" in file_type or ".tsv" in file_type:
                import csv

                dialect = csv.excel if ".csv" in file_type else csv.excel_tab
                with open(os_path.normpath(file_path), "w", newline="", encoding="utf-8") as fh:
                    csv_writer = csv.writer(fh, dialect=dialect, lineterminator="\n")
                    # Horizontal header
                    hor_header: List[str] = [""]
                    hor_header.extend(model.horizontalHeaderItem(colno).text() for colno in range(col_count))
                    csv_writer.writerow(hor_header)
                    # Vertical header + cells
                    for rowno in range(row_count):
                        row: List[str] = [model.verticalHeaderItem(rowno).text()]
                        row.extend(model.item(rowno, colno).text() for colno in range(col_count))
                        csv_writer.writerow(row)
            QMessageBox.information(self, "Success", f"The table has been successfully exported to {file_path}.")
        except PermissionError:
            QMessageBox.critical(
                self,
                "Error Exporting Cells",
                f"PermissionError: failed to export the table to {file_path}.",
            )
        else:
            model.data_exported.emit()


# https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_dialogs/wl_dialogs.py#L28
class Ng_Dialog(QDialog):
    def __init__(
        self, *args, main, title: str = "", width: int = 0, height: int = 0, resizable=True, **kwargs
    ) -> None:
        super().__init__(*args, **kwargs)
        self.main = main
        # > Dialog size
        if resizable:
            if not width:
                width = self.size().width()

            if not height:
                height = self.size().height()

            self.resize(width, height)
        else:
            if width:
                self.setFixedWidth(width)

            if height:
                self.setFixedHeight(height)
        self.setWindowTitle(title)

        # ┌———————————┐
        # │           │
        # │  content  │
        # │           │
        # │———————————│
        # │  buttons  │
        # └———————————┘
        self.content_layout = QGridLayout()
        self.button_layout = QGridLayout()

        self.grid_layout = QGridLayout()
        self.grid_layout.addLayout(self.content_layout, 0, 0)
        self.grid_layout.addLayout(self.button_layout, 1, 0)
        self.setLayout(self.grid_layout)

        # self.setSizeGripEnabled(True)

    def addWidget(self, *args, **kwargs) -> None:
        self.content_layout.addWidget(*args, **kwargs)

    def addButton(self, *args, **kwargs) -> None:
        self.button_layout.addWidget(*args, **kwargs)

    def setColumnStretch(self, column: int, strech: int) -> None:
        self.content_layout.setColumnStretch(column, strech)

    def setRowStretch(self, row: int, strech: int) -> None:
        self.content_layout.setRowStretch(row, strech)

    # Override
    def reject(self) -> None:
        pass


class Ng_Dialog_Table(Ng_Dialog):
    def __init__(
        self,
        *args,
        text: str,
        tableview: Ng_TableView,
        **kwargs,
    ) -> None:
        super().__init__(*args, **kwargs)
        self.tableview: Ng_TableView = tableview

        self.content_layout.addWidget(QLabel(text), 0, 0)
        self.content_layout.addWidget(tableview, 1, 0)

        self.button_ok = QPushButton("OK")
        self.button_ok.clicked.connect(self.accept)
        self.button_export_table = QPushButton("Export table...")
        self.button_export_table.clicked.connect(self.tableview.export_table)
        self.button_layout.addWidget(self.button_export_table, 0, 0, alignment=Qt.AlignmentFlag.AlignLeft)
        self.button_layout.addWidget(self.button_ok, 0, 1, alignment=Qt.AlignmentFlag.AlignRight)


class Ng_Worker(QObject):
    worker_done = Signal()

    def __init__(self, *args, main, dialog: Ng_Dialog, **kwargs) -> None:
        super().__init__(*args, **kwargs)
        self.main = main
        self.dialog: Ng_Dialog = dialog

    def run(self) -> None:
        raise NotImplementedError()


class Ng_Worker_SCA_Generate_Table(Ng_Worker):
    def __init__(self, *args, main, dialog: Ng_Dialog, **kwargs) -> None:
        super().__init__(*args, main=main, dialog=dialog, **kwargs)

    def run(self) -> None:
        input_file_names: Generator[str, None, None] = self.main.yield_added_file_names()
        input_file_paths: Generator[str, None, None] = self.main.yield_added_file_paths()

        sca_kwargs = {
            "is_auto_save": False,
            "odir_matched": "",
            "selected_measures": None,
            "is_reserve_parsed": self.main.checkbox_reserve_parsed_trees.isChecked(),
            "is_reserve_matched": self.main.checkbox_reserve_matched_subtrees.isChecked(),
            "is_skip_querying": False,
            "is_skip_parsing": False,
            "config": None,
        }

        attrname = "sca_analyzer"
        try:
            sca_analyzer = getattr(self.main, attrname)
        except AttributeError:
            sca_analyzer = NeoSCA(**sca_kwargs)
            setattr(self.main, attrname, sca_analyzer)
        else:
            sca_analyzer.update_options(sca_kwargs)

        err_file_paths: List[str] = []
        model: Ng_Model = self.main.model_sca
        has_trailing_rows: bool = True
        for rowno, (file_name, file_path) in enumerate(zip(input_file_names, input_file_paths)):
            try:
                counter: Optional[StructureCounter] = sca_analyzer.parse_and_query_ifile(file_path)
                # TODO should concern --no-parse, --no-query, ... after adding all available options
            except:
                err_file_paths.append(file_path)
                continue
            if counter is None:
                err_file_paths.append(file_path)
                continue
            sname_value_map: Dict[str, str] = counter.get_all_values()
            if has_trailing_rows:
                has_trailing_rows = model.removeRows(rowno, model.rowCount() - rowno)
            # Drop file_path
            _, *values = sname_value_map.values()
            model.set_row_num(rowno, values)
            model.setVerticalHeaderItem(rowno, QStandardItem(file_name))
        model.data_changed.emit()

        if err_file_paths:  # TODO: should show a table
            QMessageBox.information(
                parent=None,
                title="Error Processing Files",
                text="These files are skipped:\n- {}".format("\n- ".join(err_file_paths)),
            )
        self.worker_done.emit()


class Ng_Worker_LCA_Generate_Table(Ng_Worker):
    def __init__(self, *args, main, dialog: Ng_Dialog, **kwargs) -> None:
        super().__init__(*args, main=main, dialog=dialog, **kwargs)

    def run(self) -> None:
        input_file_names: Generator[str, None, None] = self.main.yield_added_file_names()
        input_file_paths: Generator[str, None, None] = self.main.yield_added_file_paths()

        lca_kwargs = {
            "wordlist": "bnc" if self.main.radiobutton_wordlist_BNC.isChecked() else "anc",
            "tagset": "ud" if self.main.radiobutton_tagset_ud.isChecked() else "ptb",
            "is_stdout": False,
        }
        attrname = "lca_analyzer"
        try:
            lca_analyzer = getattr(self.main, attrname)
        except AttributeError:
            lca_analyzer = LCA(**lca_kwargs)
            setattr(self.main, attrname, lca_analyzer)
        else:
            lca_analyzer.update_options(lca_kwargs)

        err_file_paths: List[str] = []
        model: Ng_Model = self.main.model_lca
        has_trailing_rows: bool = True
        for rowno, (file_name, file_path) in enumerate(zip(input_file_names, input_file_paths)):
            try:
                values = lca_analyzer._analyze(file_path=file_path)
            except:
                err_file_paths.append(file_path)
                continue
            if values is None:  # TODO: should pop up warning window
                err_file_paths.append(file_path)
                continue
            if has_trailing_rows:
                has_trailing_rows = model.removeRows(rowno, model.rowCount() - rowno)
            # Drop file_path
            del values[0]
            model.set_row_num(rowno, values)
            model.setVerticalHeaderItem(rowno, QStandardItem(file_name))
        model.data_changed.emit()

        if err_file_paths:  # TODO: should show a table
            QMessageBox.information(
                parent=None,
                title="Error Processing Files",
                text="These files are skipped:\n- {}".format("\n- ".join(err_file_paths)),
            )

        self.worker_done.emit()


class Ng_Thread(QThread):
    def __init__(self, worker: Ng_Worker):
        super().__init__()
        self.worker = worker
        # https://mayaposch.wordpress.com/2011/11/01/how-to-really-truly-use-qthreads-the-full-explanation/
        self.worker.moveToThread(self)

        # self.button_cancel = QPushButton("Cancel")
        # self.worker.dialog.addButton(self.button_cancel, 0, 0, alignment=Qt.AlignmentFlag.AlignRight)
        #
        # self.button_cancel.clicked.connect(self.cancel)
        self.started.connect(self.worker.dialog.open)
        self.finished.connect(self.worker.dialog.accept)

    def run(self):
        self.start()
        self.worker.run()

    # def cancel(self) -> None:
    #     self.terminate()
    #     self.wait()


class Ng_Main(QMainWindow):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setup_env()

        self.setWindowTitle(f"{__title__} {__version__}")
        file_path_settings = os_path.join(self.here, "ng_settings.pickle")
        self.settings_custom = SCAIO.load_pickle_file(file_path_settings, None)
        if self.settings_custom is None:
            self.settings_custom = copy.deepcopy(settings_default)
        qss = textwrap.dedent(
            f"""\
            * {{
            font-family: "{self.settings_custom['general']['ui_settings']['font_family']}";
            font-size: {self.settings_custom['general']['ui_settings']['font_size']}pt;
            }}\n"""
        )
        file_path_style_qss = os_path.join(self.here, "ng_style.qss")
        qss += Ng_QSS_Loader.read_qss_file(file_path_style_qss, "")
        self.setStyleSheet(qss)
        self.setup_menu()
        self.setup_worker()
        self.setup_main_window()

    def setup_menu(self):
        # File
        self.menu_file = QMenu("File", self.menuBar())
        action_open_file = QAction("Open File...", self.menu_file)
        action_open_file.setShortcut("CTRL+O")
        action_open_file.triggered.connect(self.menubar_file_open_file)
        action_open_folder = QAction("Open Folder...", self.menu_file)
        action_open_folder.setShortcut("CTRL+F")
        action_open_folder.triggered.connect(self.menubar_file_open_folder)
        action_restart = QAction("Restart", self.menu_file)  # TODO remove this before releasing
        action_restart.triggered.connect(self.menubar_file_restart)  # TODO remove this before releasing
        action_restart.setShortcut("CTRL+R")  # TODO remove this before releasing
        action_quit = QAction("Quit", self.menu_file)
        action_quit.setShortcut("CTRL+Q")
        action_quit.triggered.connect(self.close)
        self.menu_file.addAction(action_open_file)
        self.menu_file.addAction(action_open_folder)
        self.menu_file.addAction(action_restart)
        self.menu_file.addAction(action_quit)
        # Preferences
        self.menu_preferences = QMenu("Preferences", self.menuBar())
        action_font = QAction("Font", self.menu_preferences)
        action_font.triggered.connect(self.menubar_preferences_font)
        self.menu_preferences.addAction(action_font)
        # Help
        self.menu_help = QMenu("Help", self.menuBar())
        action_citing = QAction("Citing", self.menu_help)
        action_citing.triggered.connect(self.menubar_help_citing)
        self.menu_help.addAction(action_citing)

        self.menuBar().addMenu(self.menu_file)
        self.menuBar().addMenu(self.menu_preferences)
        self.menuBar().addMenu(self.menu_help)

    def menubar_preferences_font(self) -> None:
        ok, font = QFontDialog.getFont()
        if not ok:
            return
        breakpoint()
        print(ok, font)

    def menubar_help_citing(self) -> None:
        import json

        with open(os_path.join(self.here, "citing.json"), encoding="utf-8") as f:
            style_citation_mapping = json.load(f)
        label_citing = QLabel(f"If you use {__title__} in your research, please kindly cite as follows.")
        label_citing.setWordWrap(True)
        textedit_citing = QTextEdit()
        textedit_citing.setReadOnly(True)
        textedit_citing.setText(next(iter(style_citation_mapping.values())))
        label_choose_citation_style = QLabel("Choose citation style: ")
        combobox_choose_citation_style = QComboBox()
        combobox_choose_citation_style.addItems(tuple(style_citation_mapping.keys()))
        combobox_choose_citation_style.currentTextChanged.connect(
            lambda key: textedit_citing.setText(style_citation_mapping[key])
        )

        dialog_citing = Ng_Dialog(self, main=self, title="Citing")
        dialog_citing.addWidget(label_citing, 0, 0, 1, 2)
        dialog_citing.addWidget(label_choose_citation_style, 1, 0)
        dialog_citing.addWidget(combobox_choose_citation_style, 1, 1)
        dialog_citing.setColumnStretch(1, 1)
        dialog_citing.addWidget(textedit_citing, 2, 0, 1, 2)

        button_copy = QPushButton("Copy")
        button_copy.clicked.connect(textedit_citing.selectAll)
        button_copy.clicked.connect(textedit_citing.copy)
        button_close = QPushButton("Close")
        button_close.clicked.connect(dialog_citing.reject)

        dialog_citing.addButton(button_copy, 0, 0, alignment=Qt.AlignmentFlag.AlignLeft)
        dialog_citing.addButton(button_close, 0, 1, alignment=Qt.AlignmentFlag.AlignRight)
        dialog_citing.open()

    def setup_tab_sca(self):
        self.button_generate_table_sca = QPushButton("Generate table")
        self.button_generate_table_sca.setShortcut("CTRL+G")
        self.button_export_table_sca = QPushButton("Export all cells...")
        self.button_export_table_sca.setEnabled(False)
        # self.button_export_selected_cells = QPushButton("Export selected cells...")
        # self.button_export_selected_cells.setEnabled(False)
        self.button_clear_table_sca = QPushButton("Clear table")
        self.button_clear_table_sca.setEnabled(False)

        # TODO comment this out before releasing
        self.button_custom_func = QPushButton("Custom func")
        # TODO comment this out before releasing
        self.button_custom_func.clicked.connect(self.custom_func)

        # frame_setting_sca.setStyleSheet("background-color: pink;")
        self.checkbox_reserve_parsed_trees = QCheckBox(
            "Reserve parsed trees",
        )

        self.model_sca = Ng_Model()
        self.model_sca.setColumnCount(len(StructureCounter.DEFAULT_MEASURES))
        self.model_sca.setHorizontalHeaderLabels(StructureCounter.DEFAULT_MEASURES)
        self.model_sca.set_single_empty_row()
        self.tableview_preview_sca = Ng_TableView(main=self, model=self.model_sca)

        # Bind
        self.button_generate_table_sca.clicked.connect(self.ng_thread_sca_generate_table.start)
        self.button_export_table_sca.clicked.connect(self.tableview_preview_sca.export_table)
        self.button_clear_table_sca.clicked.connect(lambda: self.ask_clear_model(self.model_sca))
        self.model_sca.data_cleared.connect(lambda: self.button_generate_table_sca.setEnabled(True))
        self.model_sca.data_cleared.connect(lambda: self.button_export_table_sca.setEnabled(False))
        self.model_sca.data_cleared.connect(lambda: self.button_clear_table_sca.setEnabled(False))
        self.model_sca.data_changed.connect(lambda: self.button_export_table_sca.setEnabled(True))
        self.model_sca.data_changed.connect(lambda: self.button_clear_table_sca.setEnabled(True))
        self.model_sca.data_changed.connect(lambda: self.button_generate_table_sca.setEnabled(False))

        self.checkbox_reserve_parsed_trees.setChecked(True)
        self.checkbox_reserve_matched_subtrees = QCheckBox("Reserve matched subtrees")
        self.checkbox_reserve_matched_subtrees.setChecked(True)
        widget_settings_sca = QWidget()
        widget_settings_sca.setLayout(QGridLayout())
        widget_settings_sca.layout().addWidget(self.checkbox_reserve_parsed_trees, 0, 0)
        widget_settings_sca.layout().addWidget(self.checkbox_reserve_matched_subtrees, 1, 0)

        scrollarea_settings_sca = QScrollArea()
        scrollarea_settings_sca.setLayout(QGridLayout())
        scrollarea_settings_sca.setWidgetResizable(True)
        scrollarea_settings_sca.setFixedWidth(200)
        scrollarea_settings_sca.setBackgroundRole(QPalette.Light)
        scrollarea_settings_sca.setWidget(widget_settings_sca)

        self.tab_sca = QWidget()
        self.tab_sca.setLayout(QGridLayout())
        for btn_no, btn in enumerate(
            (
                self.button_generate_table_sca,
                self.button_export_table_sca,
                self.button_clear_table_sca,
                self.button_custom_func,
            ),
            start=1,
        ):
            self.tab_sca.layout().addWidget(btn, 1, btn_no - 1)
        self.tab_sca.layout().addWidget(self.tableview_preview_sca, 0, 0, 1, btn_no)
        self.tab_sca.layout().addWidget(scrollarea_settings_sca, 0, btn_no, 2, 1)
        self.tab_sca.layout().setContentsMargins(6, 4, 6, 4)

    def custom_func(self):
        breakpoint()

    def setup_tab_lca(self):
        self.button_generate_table_lca = QPushButton("Generate table")
        self.button_generate_table_lca.setShortcut("CTRL+G")
        self.button_export_table_lca = QPushButton("Export all cells...")
        self.button_export_table_lca.setEnabled(False)
        # self.button_export_selected_cells = QPushButton("Export selected cells...")
        # self.button_export_selected_cells.setEnabled(False)
        self.button_clear_table_lca = QPushButton("Clear table")
        self.button_clear_table_lca.setEnabled(False)

        self.model_lca = Ng_Model()
        self.model_lca.setColumnCount(len(LCA.FIELDNAMES) - 1)
        self.model_lca.setHorizontalHeaderLabels(LCA.FIELDNAMES[1:])
        self.model_lca.set_single_empty_row()
        self.tableview_preview_lca = Ng_TableView(main=self, model=self.model_lca)

        self.button_generate_table_lca.clicked.connect(self.ng_thread_lca_generate_table.start)
        self.button_export_table_lca.clicked.connect(self.tableview_preview_lca.export_table)
        self.button_clear_table_lca.clicked.connect(lambda: self.ask_clear_model(self.model_lca))
        self.model_lca.data_cleared.connect(lambda: self.button_generate_table_lca.setEnabled(True))
        self.model_lca.data_cleared.connect(lambda: self.button_export_table_lca.setEnabled(False))
        self.model_lca.data_cleared.connect(lambda: self.button_clear_table_lca.setEnabled(False))
        self.model_lca.data_changed.connect(lambda: self.button_export_table_lca.setEnabled(True))
        self.model_lca.data_changed.connect(lambda: self.button_clear_table_lca.setEnabled(True))
        self.model_lca.data_changed.connect(lambda: self.button_generate_table_lca.setEnabled(False))

        self.radiobutton_wordlist_BNC = QRadioButton("British National Corpus (BNC) wordlist")
        self.radiobutton_wordlist_BNC.setChecked(True)
        self.radiobutton_wordlist_ANC = QRadioButton("American National Corpus (ANC) wordlist")
        groupbox_wordlist = QGroupBox("Wordlist")
        groupbox_wordlist.setLayout(QGridLayout())
        groupbox_wordlist.layout().addWidget(self.radiobutton_wordlist_BNC, 0, 0)
        groupbox_wordlist.layout().addWidget(self.radiobutton_wordlist_ANC, 1, 0)
        self.radiobutton_tagset_ud = QRadioButton("Universal POS Tagset")
        self.radiobutton_tagset_ud.setChecked(True)
        self.radiobutton_tagset_ptb = QRadioButton("Penn Treebank POS Tagset")
        groupbox_tagset = QGroupBox("Tagset")
        groupbox_tagset.setLayout(QGridLayout())
        groupbox_tagset.layout().addWidget(self.radiobutton_tagset_ud, 0, 0)
        groupbox_tagset.layout().addWidget(self.radiobutton_tagset_ptb, 1, 0)

        widget_settings_lca = QWidget()
        widget_settings_lca.setLayout(QGridLayout())
        widget_settings_lca.layout().addWidget(groupbox_wordlist, 0, 0)
        widget_settings_lca.layout().addWidget(groupbox_tagset, 1, 0)

        scrollarea_settings_lca = QScrollArea()
        scrollarea_settings_lca.setFixedWidth(200)
        scrollarea_settings_lca.setWidgetResizable(True)
        scrollarea_settings_lca.setBackgroundRole(QPalette.Light)
        scrollarea_settings_lca.setWidget(widget_settings_lca)

        self.tab_lca = QWidget()
        self.tab_lca.setLayout(QGridLayout())

        for btn_no, btn in enumerate(
            (
                self.button_generate_table_lca,
                self.button_export_table_lca,
                self.button_clear_table_lca,
            ),
            start=1,
        ):
            self.tab_lca.layout().addWidget(btn, 1, btn_no - 1)
        self.tab_lca.layout().addWidget(self.tableview_preview_lca, 0, 0, 1, btn_no)
        self.tab_lca.layout().addWidget(scrollarea_settings_lca, 0, btn_no, 2, 1)
        self.tab_lca.layout().setContentsMargins(6, 4, 6, 4)

    def enable_button_generate_table(self, enabled: bool) -> None:
        self.button_generate_table_sca.setEnabled(enabled)
        self.button_generate_table_lca.setEnabled(enabled)

    def setup_tableview_file(self) -> None:
        self.model_file = Ng_Model()
        self.model_file.setHorizontalHeaderLabels(("Name", "Path"))
        self.model_file.data_cleared.connect(lambda: self.enable_button_generate_table(False))
        self.model_file.data_changed.connect(lambda: self.enable_button_generate_table(True))
        self.model_file.set_single_empty_row()
        self.tableview_file = Ng_TableView(main=self, model=self.model_file, has_vertical_header=False)
        self.tableview_file.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        # https://doc.qt.io/qtforpython-6/PySide6/QtWidgets/QWidget.html#PySide6.QtWidgets.PySide6.QtWidgets.QWidget.customContextMenuRequested
        self.tableview_file.customContextMenuRequested.connect(self.show_menu_for_tableview_file)

    def setup_main_window(self):
        self.setup_tab_sca()
        self.setup_tab_lca()
        self.setup_tableview_file()

        self.tabwidget = QTabWidget()
        self.tabwidget.addTab(self.tab_sca, "Syntactic Complexity Analyzer")
        self.tabwidget.addTab(self.tab_lca, "Lexical Complexity Analyzer")
        self.splitter_central_widget = QSplitter(Qt.Orientation.Vertical)
        self.splitter_central_widget.setChildrenCollapsible(False)
        self.splitter_central_widget.addWidget(self.tabwidget)
        self.splitter_central_widget.setStretchFactor(0, 2)
        self.splitter_central_widget.addWidget(self.tableview_file)
        self.splitter_central_widget.setStretchFactor(1, 1)
        self.setCentralWidget(self.splitter_central_widget)

    def setup_worker(self) -> None:
        self.dialog_processing = Ng_Dialog(
            self, main=self, title="Please Wait", width=300, height=200, resizable=False
        )
        self.label_processing = QLabel("It may take a few minutes to finish the job. Please wait.")
        self.label_processing.setWordWrap(True)
        self.dialog_processing.addWidget(self.label_processing, 0, 0, alignment=Qt.AlignmentFlag.AlignTop)

        self.ng_worker_sca_generate_table = Ng_Worker_SCA_Generate_Table(
            main=self, dialog=self.dialog_processing
        )
        self.ng_thread_sca_generate_table = Ng_Thread(self.ng_worker_sca_generate_table)

        self.ng_worker_lca_generate_table = Ng_Worker_LCA_Generate_Table(
            main=self, dialog=self.dialog_processing
        )
        self.ng_thread_lca_generate_table = Ng_Thread(self.ng_worker_lca_generate_table)

    def ask_clear_model(self, model: Ng_Model) -> None:
        if model.has_been_exported:
            model.set_single_empty_row()
        else:
            messagebox = QMessageBox(self)
            messagebox.setWindowTitle("Clear Table")
            messagebox.setText("The table has not been exported yet and all the data will be lost. Continue?")
            messagebox.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)

            messagebox.accepted.connect(model.set_single_empty_row)
            messagebox.exec()

    def setup_env(self) -> None:
        self.here = os_path.dirname(os_path.abspath(__file__))
        ng_home = os_path.dirname(self.here)
        libs_dir = os_path.join(ng_home, "libs")
        # TODO: remove these
        self.java_home = os_path.join(libs_dir, "jdk8u372")
        self.stanford_parser_home = os_path.join(libs_dir, "stanford-parser-full-2020-11-17")
        self.stanford_tregex_home = os_path.join(libs_dir, "stanford-tregex-2020-11-17")
        os.environ["JAVA_HOME"] = self.java_home
        os.environ["STANFORD_PARSER_HOME"] = self.stanford_parser_home
        os.environ["STANFORD_TREGEX_HOME"] = self.stanford_tregex_home
        self.env = os.environ.copy()

    def show_menu_for_tableview_file(self) -> None:
        menu = QMenu(self)
        action_remove_file = QAction("Remove", menu)
        action_remove_file.triggered.connect(self.remove_file_paths)
        menu.addAction(action_remove_file)
        menu.exec(QCursor.pos())

    def remove_file_paths(self) -> None:
        # https://stackoverflow.com/questions/5927499/how-to-get-selected-rows-in-qtableview
        indexes: List[QModelIndex] = self.tableview_file.selectionModel().selectedRows()
        # Remove rows from bottom to top, or otherwise the lower row indexes
        #  will change as upper rows are removed
        rownos = sorted((index.row() for index in indexes), reverse=True)
        for rowno in rownos:
            self.model_file.takeRow(rowno)
        if self.model_file.rowCount() == 0:
            self.model_file.set_single_empty_row()

    def remove_model_rows(self, model: Ng_Model, *rownos: int) -> None:
        if not rownos:
            # https://doc.qt.io/qtforpython-6/PySide6/QtGui/QStandardItemModel.html#PySide6.QtGui.PySide6.QtGui.QStandardItemModel.setRowCount
            model.setRowCount(0)
        else:
            for rowno in rownos:
                model.takeRow(rowno)

    # Type hint for generator: https://docs.python.org/3.12/library/typing.html#typing.Generator
    def yield_model_column(self, model: Ng_Model, colno: int) -> Generator[str, None, None]:
        items = (model.item(rowno, colno) for rowno in range(model.rowCount()))
        return (item.text() for item in items if item is not None)

    def yield_added_file_names(self) -> Generator[str, None, None]:
        colno_path = 0
        return self.yield_model_column(self.model_file, colno_path)

    def yield_added_file_paths(self) -> Generator[str, None, None]:
        colno_path = 1
        return self.yield_model_column(self.model_file, colno_path)

    def add_file_paths(self, file_paths_to_add: List[str]) -> None:
        unique_file_paths_to_add: Set[str] = set(file_paths_to_add)
        already_added_file_paths: Set[str] = set(self.yield_added_file_paths())
        file_paths_dup: Set[str] = unique_file_paths_to_add & already_added_file_paths
        file_paths_unsupported: Set[str] = set(
            filter(lambda p: SCAIO.suffix(p) not in SCAIO.SUPPORTED_EXTENSIONS, file_paths_to_add)
        )
        file_paths_empty: Set[str] = set(filter(lambda p: not os_path.getsize(p), unique_file_paths_to_add))
        file_paths_ok: Set[str] = (
            unique_file_paths_to_add
            - already_added_file_paths
            - file_paths_dup
            - file_paths_unsupported
            - file_paths_empty
        )
        if file_paths_ok:
            self.model_file.remove_single_empty_row()
            colno_name = 0
            already_added_file_names = list(
                self.yield_model_column(self.model_file, colno_name)
            )  # Here the already_added_file_names will have no duplicates
            for file_path in file_paths_ok:
                file_name = os_path.splitext(os_path.basename(file_path))[0]
                if file_name in already_added_file_names:
                    occurrence = 2
                    while f"{file_name} ({occurrence})" in already_added_file_names:
                        occurrence += 1
                    file_name = f"{file_name} ({occurrence})"
                already_added_file_names.append(file_name)
                rowno = self.model_file.rowCount()
                self.model_file.set_row_str(rowno, (file_name, file_path))

            self.model_file.data_changed.emit()

        if file_paths_dup or file_paths_unsupported or file_paths_empty:
            model_err_files = Ng_Model()
            model_err_files.setHorizontalHeaderLabels(("Error Type", "File Path"))
            for reason, file_paths in (
                ("Duplicate file", file_paths_dup),
                ("Unsupported file type", file_paths_unsupported),
                ("Empty file", file_paths_empty),
            ):
                for file_path in file_paths:
                    model_err_files.insertRow(
                        model_err_files.rowCount(),
                        (QStandardItem(reason), QStandardItem(file_path)),
                    )
            tableview_err_files = Ng_TableView(main=self, model=model_err_files, has_vertical_header=False)

            dialog = Ng_Dialog_Table(
                self,
                main=self,
                title="Error Adding Files",
                text="Failed to add the following files.",
                width=300,
                height=200,
                resizable=True,
                tableview=tableview_err_files,
            )
            dialog.open()

    def menubar_file_open_folder(self):
        # TODO: Currently only include files of supported types, should include
        #  all files, and popup error for unsupported files
        folder_dialog = QFileDialog(self)
        # TODO remove default directory before releasing
        folder_path = folder_dialog.getExistingDirectory(
            caption="Open Folder",
            dir='directory="/home/tan/docx/corpus/YuHua-parallel-corpus-zh-en/02aligned/standalone/',
        )
        if not folder_path:
            return

        file_paths_to_add = []
        for extension in SCAIO.SUPPORTED_EXTENSIONS:
            file_paths_to_add.extend(glob.glob(os_path.join(folder_path, f"*.{extension}")))
        self.add_file_paths(file_paths_to_add)

    def menubar_file_open_file(self):
        file_dialog = QFileDialog(self)
        file_paths_to_add, _ = file_dialog.getOpenFileNames(
            parent=None,
            caption="Open Files",
            dir="/home/tan/docx/corpus/YuHua-parallel-corpus-zh-en/02aligned/standalone/",
            # TODO remove this before releasing
            filter="Text files (*.txt);;Docx files (*.docx);;Odt files (*.odt);;All files (*.*)",
        )
        if not file_paths_to_add:
            return
        self.add_file_paths(file_paths_to_add)

    def menubar_file_restart(self):
        self.close()
        command = [sys.executable, "-m", "neosca_gui"]
        subprocess.call(command, env=os.environ.copy(), close_fds=False)


def main():
    ui_scaling = DEFAULT_INTERFACE_SCALING
    os.environ["QT_SCALE_FACTOR"] = re.sub(r"([0-9]{2})%$", r".\1", ui_scaling)
    ng_app = QApplication(sys.argv)
    ng_window = Ng_Main()
    ng_window.showMaximized()
    sys.exit(ng_app.exec())
